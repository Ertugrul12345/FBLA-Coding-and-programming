import customtkinter as ctk
import mysql.connector
from tkinter import messagebox, IntVar, Text, Toplevel
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import nltk
from nltk.chat.util import Chat, reflections
from tkinter.scrolledtext import ScrolledText
import webbrowser
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
checkbox_states = {} 
ctk.set_appearance_mode("dark")  
ctk.set_default_color_theme("blue") 

login_success = False

pairs = [
    {"question": "add new partner", "answers": ["To add a new partner, navigate to the 'Add Partner' section from the main menu, then fill out the necessary information and click 'Add Partner'."]},
    {"question": "send report", "answers": ["To send a report  click on the 'Generate Report' button located in the main window. Then, enter your email address in the prompted window and click 'send' to receive the report."]}

]

def nltk_chatbot_response(question):
    chat = Chat(pairs, reflections)
    return chat.respond(question)


def login():
    global login_success
    login_success = False 
    app = ctk.CTk()
    app.geometry("600x440")
    app.title('Login')

    def button_function():
        global login_success
        username = entry1.get()
        password = entry2.get()
        if username == "user@gmail.com" and password == "user":
            messagebox.showinfo("Login Success", "You have successfully logged in")
            login_success = True  
            app.destroy()
        else:
            messagebox.showinfo("Login Failed", "Invalid username or password")

    frame = ctk.CTkFrame(master=app, width=320, height=360, corner_radius=15)
    frame.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)

    l2 = ctk.CTkLabel(master=frame, text="Log into your Account", font=('Century Gothic', 20))
    l2.place(x=50, y=45)

    entry1 = ctk.CTkEntry(master=frame, width=220, placeholder_text='User ID')
    entry1.place(x=50, y=110)

    entry2 = ctk.CTkEntry(master=frame, width=220, placeholder_text='Password', show="*")
    entry2.place(x=50, y=165)

    button1 = ctk.CTkButton(master=frame, width=220, text="Login", command=button_function, corner_radius=6)
    button1.place(x=50, y=220)

    app.protocol("WM_DELETE_WINDOW", app.destroy)  

    app.mainloop()

    return login_success  


DB_HOST = 'localhost'
DB_USER = 'root'
DB_PASSWORD = 'Ekkomekko.25'
DB_NAME = 'testdb'

def logout():
    # Close the main window
    app.destroy()
    
    # Call the login function to display the login page again
    login()

def add_partner_to_db(name, org_type, location, contact, window, resources):
    try:
        connection = mysql.connector.connect(
            host=DB_HOST,
            user=DB_USER,
            password=DB_PASSWORD,
            database=DB_NAME
        )
        cursor = connection.cursor()
        
        query = """
                    INSERT INTO companies (name, phone_number, business_type, location, resources) 
                    VALUES (%s, %s, %s, %s, %s)
                """
        values = (name, contact, org_type, location, resources) 

        cursor.execute(query, values)
        connection.commit()
        
        messagebox.showinfo("Success", "Partner added successfully to the database.")
        window.destroy()
    
    except mysql.connector.Error as e:
        messagebox.showerror("Error", f"Error connecting to MySQL database: {e}")
    
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()



def add_partner_window():
    add_window = ctk.CTkToplevel(app)  
    add_window.title("Add New Partner")
    add_window.geometry("400x400")  

    entry_width = 250
    label_font = ('Roboto', 14)
    entry_font = ('Roboto', 12)
    button_font = ('Roboto', 14)
    padding_x = 20 
    padding_y = 10

    def create_labeled_entry(label_text, placeholder):
        frame = ctk.CTkFrame(add_window)
        frame.pack(padx=padding_x, pady=padding_y, fill='x')
        
        label = ctk.CTkLabel(frame, text=label_text, font=label_font, width=100, anchor='w')
        label.pack(side='left')
        
        entry = ctk.CTkEntry(frame, placeholder_text=placeholder, width=entry_width, font=entry_font)
        entry.pack(side='left', padx=(10, 0))
        
        return entry


    entry_name = create_labeled_entry("Name:", "FBLA")
    entry_type = create_labeled_entry("Type:", "Business")
    entry_location = create_labeled_entry("Location:", "North Carolina")
    entry_contact = create_labeled_entry("Contact:", "+123456789")
    entry_resources = create_labeled_entry("Resources:", "Enter available resources") 


    button_add = ctk.CTkButton(add_window, text="Add Partner", width=120, height=40, font=button_font,
                               command=lambda: add_partner_to_db(
                                   entry_name.get(),
                                   entry_type.get(),
                                   entry_location.get(),
                                   entry_contact.get(),
                                   add_window,
                                   entry_resources.get()
                               ))
    button_add.pack(pady=20)  
    add_window.mainloop()



import spacy

nlp = spacy.load("en_core_web_md")  


qa_pairs = [{"question": nlp(pair["question"]), "answers": pair["answers"]} for pair in pairs]

def get_best_response(user_input):
    user_doc = nlp(user_input)
    best_score = 0
    best_answer = "I'm not sure how to help with that. Can you try asking in a different way?"
    
    for pair in qa_pairs:
        score = user_doc.similarity(pair["question"])
        if score > best_score:
            best_score = score
            best_answer = pair["answers"][0]
    
    return best_answer

def open_help_window(parent_window):
    help_window = ctk.CTkToplevel(parent_window)
    help_window.title("Help")
    help_window.geometry("500x400")
    ctk.set_appearance_mode("dark")

    chat_frame = ctk.CTkFrame(help_window, corner_radius=10)
    chat_frame.pack(padx=10, pady=10, fill='both', expand=True)

    chat_text_area = Text(chat_frame, wrap='word', font=('Arial', 12), bg='#2e2e2e', fg='#ffffff', state='disabled', height=15)
    chat_text_area.pack(padx=5, pady=5, fill='both', expand=True)

    input_frame = ctk.CTkFrame(help_window, corner_radius=10)
    input_frame.pack(padx=10, pady=10, fill='x')

    help_entry = ctk.CTkEntry(input_frame, placeholder_text="Type your question here...")
    help_entry.pack(side='left', padx=(0, 10), fill='x', expand=True)

    def send_question():
        user_question = help_entry.get()
        if user_question:
            response = get_best_response(user_question)
            chat_text_area.config(state='normal')
            chat_text_area.insert('end', f"User: {user_question}\nAI: {response}\n")
            chat_text_area.config(state='disabled')
            help_entry.delete(0, 'end')
            chat_text_area.see('end')
        else:
            messagebox.showinfo("Empty Input", "Please enter a question.", parent=help_window)

    send_button = ctk.CTkButton(input_frame, text="Send", command=send_question)
    send_button.pack(side='left', padx=10)

    def open_github_and_close_help():
        webbrowser.open_new("https://github.com/ERTUGRULFND/FBLA-PartnerPedia/blob/main/README.md")
        help_window.destroy()

    github_button = ctk.CTkButton(help_window, text="More Help ", command=open_github_and_close_help)
    github_button.pack(pady=10)

    help_window.attributes('-topmost', True)
    help_window.focus_force()
    
def on_close_help_window(parent_window):
    parent_window.deiconify()


def fetch_data(search_term, search_filter, Partner_Frame):
    global checkbox_states
    checkbox_states.clear()  
    for widget in Partner_Frame.winfo_children():
        widget.destroy()

    try:
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='Ekkomekko.25',
            database='testdb'
        )
        if connection.is_connected():
            cursor = connection.cursor()
    
            query = """SELECT name, phone_number, business_type, location, resources
                       FROM companies
                       WHERE name LIKE %s OR phone_number LIKE %s OR business_type LIKE %s OR location LIKE %s OR resources LIKE %s"""
            search_value = "%" + search_term + "%"
            cursor.execute(query, (search_value, search_value, search_value, search_value, search_value))

            records = cursor.fetchall()

          
            label_widths = [200, 120, 150, 150, 200] 

            for index, row in enumerate(records):
                partner_frame = ctk.CTkFrame(Partner_Frame)
                partner_frame.pack(pady=2, fill='x', padx=10, expand=True)

                company_identifier = row[0]
                chk_value = IntVar()

                checkbox = ctk.CTkCheckBox(partner_frame, variable=chk_value, text="")
                checkbox.pack(side='left', padx=10)

                for i, (text, width) in enumerate(zip(row, label_widths)):
                    label = ctk.CTkLabel(partner_frame, text=text, width=width, anchor="w")
                    label.pack(side='left', padx=5, pady=5)

                if index < len(records) - 1:
                    separator = ctk.CTkFrame(Partner_Frame, height=2, bg_color="#D1D1D1")
                    separator.pack(fill='x', padx=10, pady=5)

                checkbox_states[company_identifier] = (chk_value, partner_frame)

            cursor.close()
        connection.close()
    except mysql.connector.Error as e:
        messagebox.showerror("Database Connection Error", f"Error connecting to MySQL: {e}")



def delete_selected_companies(Partner_Frame):
    global checkbox_states
    companies_to_delete = [name for name, (var, _) in checkbox_states.items() if var.get() == 1]
    if not companies_to_delete:
        messagebox.showinfo("Delete Companies", "No companies selected for deletion.")
        return
    
    try:
        connection = mysql.connector.connect(host='localhost', user='root', password='Ekkomekko.25', database='testdb')
        cursor = connection.cursor()
        for company_name in companies_to_delete:
            query = "DELETE FROM companies WHERE name = %s"
            cursor.execute(query, (company_name,))
        connection.commit()
        messagebox.showinfo("Delete Companies", "Selected companies have been deleted.")
    except mysql.connector.Error as e:
        messagebox.showerror("Database Connection Error", f"Error connecting to MySQL: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
    
    fetch_data("", "All", Partner_Frame)  
    for company_name in companies_to_delete:
        try:
            elements = checkbox_states.get(company_name, [])
            frame_to_remove = elements[1] if len(elements) > 1 else None
            separator_to_remove = elements[2] if len(elements) > 2 else None
            
            if frame_to_remove:
                frame_to_remove.destroy()
            if separator_to_remove:  
                separator_to_remove.destroy()
                

            del checkbox_states[company_name]
        except Exception as e:
            print(f"Error removing UI elements for {company_name}: {e}")






def view_partners_window(root):
    root.withdraw()

    partner_window = ctk.CTkToplevel()
    partner_window.title('View Partners')
    partner_window.geometry('1150x600')  

    screen_width = partner_window.winfo_screenwidth()
    screen_height = partner_window.winfo_screenheight()
    x_position = (screen_width - 1150) // 2
    y_position = (screen_height - 600) // 2
    partner_window.geometry(f"1150x600+{x_position}+{y_position}")

    search_frame = ctk.CTkFrame(partner_window)
    search_frame.pack(fill='x', padx=20, pady=10)


    search_filter = ctk.CTkComboBox(search_frame, values=["All", "Name", "Phone Number", "Type", "Location", "Resources"], width=120)
    search_filter.pack(side='left', padx=(0, 10))
    search_filter.set("All")

    search_entry = ctk.CTkEntry(search_frame, width=400)
    search_entry.pack(side='left', fill='x', expand=True, padx=(0, 10))


    Partner_Frame = ctk.CTkScrollableFrame(partner_window, width=1300, height=400)
    Partner_Frame.pack(padx=20, pady=(10, 0))

    search_button = ctk.CTkButton(search_frame, text="Search", command=lambda: fetch_data(search_entry.get(), search_filter.get(), Partner_Frame))
    search_button.pack(side='right')

    buttons_frame = ctk.CTkFrame(partner_window)
    buttons_frame.pack(pady=20, expand=True)

    button_font = ('bold', 19)
    back_button = ctk.CTkButton(buttons_frame, text="Back", width=150, height=50, font=button_font, command=lambda: [partner_window.destroy(), root.deiconify()])
    delete_button = ctk.CTkButton(buttons_frame, text="Delete", width=150, height=50, font=button_font, command=lambda: delete_selected_companies(Partner_Frame))
    help_button = ctk.CTkButton(buttons_frame, text="Help", width=150, height=50, font=button_font)

    back_button.pack(side='left', padx=5)
    delete_button.pack(side='left', padx=5)
    help_button.pack(side='left', padx=5)

    buttons_frame.pack_configure(anchor='center')

    fetch_data("", "All", Partner_Frame)





def on_close_partner_window(root):
    root.deiconify()

def send_report_window(parent_window):
    email_window = ctk.CTkToplevel(parent_window)
    email_window.title("Send Report")
    email_window.geometry("500x250")

    frame = ctk.CTkFrame(email_window, corner_radius=10)
    frame.pack(padx=30, pady=30, fill='both', expand=True)

    label = ctk.CTkLabel(frame, text="Enter your Gmail to Receive Report:", font=('bold', 14))
    label.pack(pady=(0, 20))

    email_entry = ctk.CTkEntry(frame, width=400, height=40, font=('bold', 14), placeholder_text="example@gmail.com")
    email_entry.pack(pady=(0, 20))

    send_button = ctk.CTkButton(frame, text="Send Report", command=lambda: send_email_report(email_entry.get()), fg_color="#0078D7", text_color="white", hover_color="#0053a4")
    send_button.pack()



def main_window():
    global app
    app = ctk.CTk()
    app.title('PartnerPedia')
    app.geometry('700x550')


    window_width = 700
    window_height = 550

    screen_width = app.winfo_screenwidth()
    screen_height = app.winfo_screenheight()

    x = screen_width - window_width  
    y = (screen_height - window_height) // 2  

    app.geometry(f"{window_width}x{window_height}+{x}+{y}")

    ctk.set_appearance_mode("dark")  
    ctk.set_default_color_theme("blue")  

    welcome_label = ctk.CTkLabel(app, text="Welcome to PartnerPedia!", font=('Roboto Medium', 24),
                                anchor="center", width=780, height=60)
    welcome_label.grid(row=0, column=0, padx=10, pady=(10, 20), sticky='ew')

    button_font = ('Roboto Medium', 20)

    view_button = ctk.CTkButton(app, text="View Partners", width=300, height=60, corner_radius=10,
                                font=button_font, fg_color="#007FFF", hover_color="#005FCC",
                                command=lambda: view_partners_window(app))
    view_button.grid(row=1, column=0, sticky='ew', padx=50, pady=10)

    add_partner_button = ctk.CTkButton(app, text="Add Partner", width=300, height=60, corner_radius=10,
                                       font=button_font, fg_color="#007FFF", hover_color="#005FCC",
                                       command=add_partner_window)
    add_partner_button.grid(row=2, column=0, sticky='ew', padx=50, pady=10)

    send_report_button = ctk.CTkButton(app, text="Send Report", width=300, height=60, corner_radius=10,
                                    font=button_font, fg_color="#007FFF", hover_color="#005FCC",
                                    command=lambda: send_report_window(app))

    send_report_button.grid(row=3, column=0, sticky='ew', padx=50, pady=10)

    help_button = ctk.CTkButton(app, text="Help", width=300, height=60, corner_radius=10,
                                font=button_font, fg_color="#007FFF", hover_color="#005FCC",
                                command=lambda: open_help_window(app))
    help_button.grid(row=4, column=0, sticky='ew', padx=50, pady=10)

    logout_button = ctk.CTkButton(app, text="Log Out", width=300, height=60, corner_radius=10,
                                  font=button_font, fg_color="#FF6347", hover_color="#FF4500",
                                  command=logout)
    logout_button.grid(row=5, column=0, sticky='ew', padx=50, pady=10)

    app.grid_columnconfigure(0, weight=1)
    app.mainloop()




import openpyxl
from openpyxl.styles import Font, Alignment

def generate_excel_report(data, filename="Business_Partners_Report.xlsx"):
    """
    Generate an Excel report with partner information.
    
    Parameters:
    - data: A list of tuples containing the partner data to be written to the Excel file.
            Each tuple represents a row in the Excel file and should follow the format:
            (Name, Phone Number, Business Type, Location)
    - filename: The name of the Excel file to be saved (default is 'Business_Partners_Report.xlsx').
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Partners"
    
    headers = ["Name", "Phone Number", "Business Type", "Location"]
    sheet.append(headers) 
    
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    for row_data in data:
        sheet.append(row_data)
    
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2 
    
    workbook.save(filename)



def send_email_report(recipient_email):
    if not recipient_email:
        messagebox.showerror("Error", "Please enter a valid email address.")
        return

    try:
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='Ekkomekko.25',
            database='testdb'
        )
        cursor = connection.cursor()
        query = "SELECT name, phone_number, business_type, location FROM companies"
        cursor.execute(query)
        records = cursor.fetchall()
        
        report_filename = "Business_Partners_Report.xlsx"
        generate_excel_report(records, filename=report_filename)
            
    except mysql.connector.Error as e:
        messagebox.showerror("Database Connection Error", f"Error connecting to MySQL: {e}")
        return
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
    
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = "partnerpediaa@gmail.com" 
    sender_password = "hpsm wlga spab ljzv"  
    
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = recipient_email
    message["Subject"] = "Business Partners Report"
    
    with open(report_filename, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {os.path.basename(report_filename)}",
    )
    message.attach(part)
    
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(message)
        server.quit()
        messagebox.showinfo("Success", "Report sent successfully.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to send email: {e}")
    finally:
        if os.path.exists(report_filename):
            os.remove(report_filename)  

if __name__ == "__main__":
    login_success = login()  
    if login_success:  
        main_window()  

         
